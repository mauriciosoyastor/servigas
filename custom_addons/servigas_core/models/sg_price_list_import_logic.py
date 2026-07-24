"""Lógica pura (sin ORM) para importar listas de precios Servigas."""

from __future__ import annotations

import csv
import io
import re
from typing import Any

REJECTED_EXTENSIONS = {
    ".pdf",
    ".png",
    ".jpg",
    ".jpeg",
    ".gif",
    ".webp",
    ".tif",
    ".tiff",
    ".bmp",
}

FIELD_ALIASES = {
    "barcode": {
        "barcode",
        "codigo de barras",
        "código de barras",
        "ean",
        "codigo barras",
    },
    "default_code": {
        "default code",
        "default_code",
        "codigo",
        "código",
        "codigo interno",
        "código interno",
        "referencia",
        "sku",
        "codigo fabricante",
        "código fabricante",
    },
    "name": {
        "name",
        "nombre",
        "descripcion",
        "descripción",
        "producto",
        "detalle",
    },
    "list_price": {
        "list price",
        "list_price",
        "precio",
        "precio venta",
        "precio de venta",
        "precio publico",
        "precio público",
        "pvp",
    },
    "standard_price": {
        "standard price",
        "standard_price",
        "costo",
        "costo unitario",
        "precio costo",
        "cost",
    },
}

TEMPLATE_CSV = (
    "barcode,default_code,name,list_price,standard_price\n"
    "7790000000000,SKU-EJEMPLO,Producto ejemplo,1500.00,900.00\n"
)


def is_rejected_filename(filename: str | None) -> bool:
    if not filename:
        return False
    name = filename.strip().lower()
    return any(name.endswith(ext) for ext in REJECTED_EXTENSIONS)


def _norm_header(value: str) -> str:
    text = (value or "").strip().lower()
    text = text.replace("_", " ")
    text = re.sub(r"\s+", " ", text)
    return text


def suggest_mapping(headers: list[str]) -> dict[str, str]:
    mapping: dict[str, str] = {}
    used_headers: set[str] = set()
    normalized = {h: _norm_header(h) for h in headers if h is not None}
    for field, aliases in FIELD_ALIASES.items():
        for header, norm in normalized.items():
            if header in used_headers:
                continue
            if norm in aliases:
                mapping[field] = header
                used_headers.add(header)
                break
    return mapping


def parse_price(value: Any) -> tuple[float | None, bool]:
    """Return (amount_or_None_if_empty, is_invalid)."""
    if value is None:
        return None, False
    text = str(value).strip()
    if text == "":
        return None, False
    cleaned = text.replace("$", "").replace(" ", "")
    if "," in cleaned and "." in cleaned:
        if cleaned.rfind(",") > cleaned.rfind("."):
            cleaned = cleaned.replace(".", "").replace(",", ".")
        else:
            cleaned = cleaned.replace(",", "")
    elif "," in cleaned:
        cleaned = cleaned.replace(",", ".")
    try:
        amount = float(cleaned)
    except ValueError:
        return None, True
    if amount < 0:
        return None, True
    return amount, False


def normalize_row(raw: dict[str, Any], mapping: dict[str, str]) -> dict[str, Any]:
    def cell(field: str) -> str:
        header = mapping.get(field)
        if not header:
            return ""
        return str(raw.get(header, "") or "").strip()

    price_errors: list[str] = []
    list_price, list_invalid = parse_price(cell("list_price") if mapping.get("list_price") else "")
    if mapping.get("list_price") and list_invalid:
        price_errors.append("list_price")
    standard_price, cost_invalid = parse_price(
        cell("standard_price") if mapping.get("standard_price") else ""
    )
    if mapping.get("standard_price") and cost_invalid:
        price_errors.append("standard_price")

    return {
        "barcode": cell("barcode"),
        "default_code": cell("default_code"),
        "name": cell("name"),
        "list_price": list_price,
        "standard_price": standard_price,
        "price_errors": price_errors,
    }


def match_product(row: dict[str, Any], indexes: dict[str, dict[str, list[int]]]) -> dict[str, Any]:
    if row.get("price_errors"):
        return {"status": "error", "product_id": None, "candidates": [], "reason": "invalid_price"}
    name = (row.get("name") or "").strip()
    if not name:
        return {"status": "error", "product_id": None, "candidates": [], "reason": "missing_name"}

    barcode = (row.get("barcode") or "").strip()
    code = (row.get("default_code") or "").strip()
    by_barcode = indexes.get("by_barcode") or {}
    by_code = indexes.get("by_code") or {}
    by_name = indexes.get("by_name") or {}

    if barcode:
        hits = by_barcode.get(barcode, [])
        if len(hits) == 1:
            return {"status": "update", "product_id": hits[0], "candidates": [], "reason": "barcode"}
        if len(hits) > 1:
            return {
                "status": "review",
                "product_id": None,
                "candidates": list(hits),
                "reason": "ambiguous_barcode",
            }

    if code:
        hits = by_code.get(code, [])
        if len(hits) == 1:
            return {"status": "update", "product_id": hits[0], "candidates": [], "reason": "default_code"}
        if len(hits) > 1:
            return {
                "status": "review",
                "product_id": None,
                "candidates": list(hits),
                "reason": "ambiguous_code",
            }

    name_key = name.casefold()
    hits = by_name.get(name_key, [])
    # Also try exact upper key used by indexes builders
    if not hits:
        hits = by_name.get(name.upper(), [])
    if len(hits) == 1:
        return {"status": "update", "product_id": hits[0], "candidates": [], "reason": "name"}
    if len(hits) > 1:
        return {
            "status": "review",
            "product_id": None,
            "candidates": list(hits),
            "reason": "ambiguous_name",
        }

    return {"status": "create", "product_id": None, "candidates": [], "reason": "no_match"}


def classify_rows(
    raw_rows: list[dict[str, Any]],
    mapping: dict[str, str],
    indexes: dict[str, dict[str, list[int]]],
) -> list[dict[str, Any]]:
    classified: list[dict[str, Any]] = []
    for index, raw in enumerate(raw_rows, start=1):
        normalized = normalize_row(raw, mapping)
        matched = match_product(normalized, indexes)
        classified.append(
            {
                "line_number": index,
                **normalized,
                **matched,
            }
        )
    return classified


def _parse_csv_bytes(raw_bytes: bytes) -> dict[str, Any]:
    text = raw_bytes.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        return {"headers": [], "rows": [], "error": "El archivo CSV no tiene encabezados."}
    headers = [h for h in reader.fieldnames if h is not None]
    rows = []
    for row in reader:
        rows.append({k: (v if v is not None else "") for k, v in row.items() if k is not None})
    return {"headers": headers, "rows": rows, "error": None}


def _parse_xlsx_bytes(raw_bytes: bytes) -> dict[str, Any]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {
            "headers": [],
            "rows": [],
            "error": "Para importar Excel (.xlsx) hace falta openpyxl en el servidor. "
            "Exportá a CSV o instalá openpyxl.",
        }
    workbook = load_workbook(filename=io.BytesIO(raw_bytes), read_only=True, data_only=True)
    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return {"headers": [], "rows": [], "error": "El Excel está vacío."}
    headers = [str(c).strip() if c is not None else "" for c in header_row]
    if not any(headers):
        return {"headers": [], "rows": [], "error": "El Excel no tiene encabezados."}
    # Keep original header labels; blank headers get placeholders
    clean_headers = []
    for i, h in enumerate(headers):
        clean_headers.append(h if h else f"col_{i + 1}")
    rows = []
    for values in rows_iter:
        if values is None or all(v is None or str(v).strip() == "" for v in values):
            continue
        item = {}
        for i, header in enumerate(clean_headers):
            value = values[i] if i < len(values) else ""
            item[header] = "" if value is None else value
        rows.append(item)
    return {"headers": clean_headers, "rows": rows, "error": None}


def parse_tabular_bytes(filename: str | None, raw_bytes: bytes) -> dict[str, Any]:
    if is_rejected_filename(filename):
        return {
            "headers": [],
            "rows": [],
            "error": "PDF e imágenes no se importan en esta versión. Convertí la lista a Excel o CSV.",
        }
    name = (filename or "").strip().lower()
    if name.endswith(".csv") or name.endswith(".txt"):
        return _parse_csv_bytes(raw_bytes)
    if name.endswith(".xlsx"):
        return _parse_xlsx_bytes(raw_bytes)
    if name.endswith(".xls"):
        return {
            "headers": [],
            "rows": [],
            "error": "El formato .xls antiguo no está soportado. Guardá como .xlsx o CSV.",
        }
    # Heurística: intentar CSV
    return _parse_csv_bytes(raw_bytes)


def build_product_indexes(products: list[dict[str, Any]]) -> dict[str, dict[str, list[int]]]:
    """products: list of {id, barcode, default_code, name}."""
    by_barcode: dict[str, list[int]] = {}
    by_code: dict[str, list[int]] = {}
    by_name: dict[str, list[int]] = {}
    for product in products:
        pid = product["id"]
        barcode = (product.get("barcode") or "").strip()
        code = (product.get("default_code") or "").strip()
        name = (product.get("name") or "").strip()
        if barcode:
            by_barcode.setdefault(barcode, []).append(pid)
        if code:
            by_code.setdefault(code, []).append(pid)
        if name:
            by_name.setdefault(name.casefold(), []).append(pid)
            by_name.setdefault(name.upper(), []).append(pid)
    return {"by_barcode": by_barcode, "by_code": by_code, "by_name": by_name}
