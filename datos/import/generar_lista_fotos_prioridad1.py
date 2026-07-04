"""Genera checklist de fotos para marcas prioridad 1 (Eskabe, Longvie, Brogas, Fercor, Orbis)."""

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

from imagenes_producto import find_image_path

IMPORT_DIR = Path(__file__).parent
INPUT_CSV = IMPORT_DIR / "maestro_import_odoo_final.csv"
OUTPUT_CSV = IMPORT_DIR / "lista_fotos_prioridad1.csv"
OUTPUT_XLSX = IMPORT_DIR / "lista_fotos_prioridad1.xlsx"

PRIORIDAD_1_MARCAS = {"Eskabe", "Longvie", "Brogas", "Orbis"}
PRIORIDAD_1_PROVEEDORES = {"Fercor"}


def is_orbis(row: dict) -> bool:
    marca = (row.get("marca") or "").strip().lower()
    descripcion = (row.get("descripcion") or "").lower()
    return marca == "orbis" or "orbis" in descripcion


def marca_prioridad(row: dict) -> str | None:
    if is_orbis(row):
        return "Orbis"
    marca = (row.get("marca") or "").strip()
    proveedor = (row.get("proveedor") or "").strip()
    if marca in PRIORIDAD_1_MARCAS:
        return marca
    if proveedor in PRIORIDAD_1_PROVEEDORES:
        return "Fercor"
    return None


def is_prioridad_1(row: dict) -> bool:
    return marca_prioridad(row) is not None


def style_sheet(path: Path, sheet_name: str) -> None:
    wb = load_workbook(path)
    ws = wb[sheet_name]
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(path)


def main() -> None:
    df = pd.read_csv(INPUT_CSV, encoding="utf-8-sig")
    df["marca_prioridad"] = df.apply(marca_prioridad, axis=1)
    filtered = df[df["marca_prioridad"].notna()].copy()

    filtered["archivo_imagen"] = filtered["codigo_fabricante"].astype(str) + ".webp"
    filtered["tiene_foto"] = filtered["codigo_fabricante"].astype(str).apply(
        lambda code: "si" if find_image_path(code) else "no"
    )

    columns = [
        "codigo_fabricante",
        "descripcion",
        "marca_prioridad",
        "marca",
        "categoria_producto",
        "proveedor",
        "archivo_imagen",
        "tiene_foto",
        "precio_publico",
        "stock_deposito",
        "notas",
    ]
    result = filtered[columns].sort_values(
        ["marca_prioridad", "categoria_producto", "codigo_fabricante"]
    )

    result.to_csv(OUTPUT_CSV, index=False, encoding="utf-8-sig")

    with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as writer:
        result.to_excel(writer, sheet_name="Prioridad 1", index=False)
        for marca in sorted(result["marca_prioridad"].unique()):
            result[result["marca_prioridad"] == marca].to_excel(
                writer, sheet_name=marca[:31], index=False
            )

    style_sheet(OUTPUT_XLSX, "Prioridad 1")
    for marca in sorted(result["marca_prioridad"].unique()):
        style_sheet(OUTPUT_XLSX, marca[:31])

    total = len(result)
    con_foto = (result["tiene_foto"] == "si").sum()
    print(f"Generado: {OUTPUT_CSV}")
    print(f"Generado: {OUTPUT_XLSX}")
    print(f"Productos prioridad 1: {total}")
    print(f"Con foto ya cargada: {con_foto}")
    print("Por marca_prioridad:")
    for marca, count in result["marca_prioridad"].value_counts().sort_index().items():
        print(f"  {marca}: {count}")


if __name__ == "__main__":
    main()
