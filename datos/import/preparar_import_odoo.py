"""Genera maestro final listo para importar a Odoo (sin descuentos automáticos)."""

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

INPUT = Path(__file__).parent / "maestro_productos_consolidado.xlsx"
OUTPUT_XLSX = Path(__file__).parent / "maestro_import_odoo_final.xlsx"
OUTPUT_CSV = Path(__file__).parent / "maestro_import_odoo_final.csv"


def style_header(path: Path) -> None:
    wb = load_workbook(path)
    ws = wb["Productos"]
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(path)


def main() -> None:
    df = pd.read_excel(INPUT, sheet_name="Productos")

    final = pd.DataFrame(
        {
            "codigo_fabricante": df["codigo_fabricante"],
            "descripcion": df["descripcion"],
            "marca": df["marca"],
            "categoria_producto": df["categoria_producto"],
            "unidad_medida": df["unidad_medida"].fillna("Unidades"),
            "precio_publico": df["precio_publico"],
            "costo": df["costo"].where(df["costo"].notna(), df["precio_publico"]),
            "proveedor": df["proveedor"],
            "codigo_barras": df["codigo_barras"].where(
                df["codigo_barras"].notna(), df["codigo_fabricante"]
            ),
            "stock_deposito": pd.to_numeric(df["stock_deposito"], errors="coerce"),
            "activo": df["activo"].fillna("si"),
            "notas": df["notas"].fillna(""),
            "fuente_archivo": df["fuente_archivo"],
        }
    )

    final["stock_deposito"] = final["stock_deposito"].fillna(0)
    final["activo"] = final["activo"].astype(str).str.lower().replace({"true": "si", "false": "no"})

    instrucciones = pd.DataFrame(
        {
            "tema": [
                "Alcance",
                "Precios",
                "Descuentos",
                "Stock",
                "Import Odoo",
                "POS",
            ],
            "detalle": [
                "Se importan los 8767 productos de todas las fuentes sin filtrar.",
                "Un solo precio por producto (precio_publico → list_price en Odoo).",
                "Sin listas instalador/mayorista. El vendedor aplicará descuentos manualmente en POS.",
                "Solo 77 productos traen stock del Excel Fercor. El resto inicia en 0.",
                "Usar script import_a_odoo.py o Inventario → Productos → Importar este archivo.",
                "Configurar POS con permiso de descuento manual por línea o pedido.",
            ],
        }
    )

    resumen = pd.DataFrame(
        [
            {"metrica": "Total productos", "valor": len(final)},
            {"metrica": "Con stock > 0", "valor": int((final["stock_deposito"] > 0).sum())},
            {"metrica": "Sin precio", "valor": int(final["precio_publico"].isna().sum())},
            {"metrica": "Marcas distintas", "valor": final["marca"].nunique()},
            {"metrica": "Proveedores distintos", "valor": final["proveedor"].nunique()},
        ]
    )

    with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as writer:
        final.to_excel(writer, sheet_name="Productos", index=False)
        instrucciones.to_excel(writer, sheet_name="Instrucciones", index=False)
        resumen.to_excel(writer, sheet_name="Resumen", index=False)

    style_header(OUTPUT_XLSX)

    export_csv = final.drop(columns=["fuente_archivo"])
    export_csv.to_csv(OUTPUT_CSV, index=False, encoding="utf-8-sig")

    print(f"Productos: {len(final)}")
    print(f"Con stock: {(final['stock_deposito'] > 0).sum()}")
    print(f"Generado: {OUTPUT_XLSX}")
    print(f"Generado: {OUTPUT_CSV}")


if __name__ == "__main__":
    main()
