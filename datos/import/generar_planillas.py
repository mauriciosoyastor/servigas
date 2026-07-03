"""Genera planillas Excel para importación Odoo y puente Factura Web."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUTPUT_DIR = Path(__file__).parent

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
INSTR_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")


def style_header(ws, row: int, col_count: int) -> None:
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def autosize_columns(ws, min_width: int = 12, max_width: int = 40) -> None:
    for col_cells in ws.columns:
        letter = get_column_letter(col_cells[0].column)
        length = max(len(str(cell.value or "")) for cell in col_cells)
        ws.column_dimensions[letter].width = min(max(length + 2, min_width), max_width)


def create_maestro_productos() -> None:
    wb = Workbook()

    # --- Hoja datos ---
    ws = wb.active
    ws.title = "Productos"

    headers = [
        "codigo_fabricante",
        "descripcion",
        "marca",
        "categoria_producto",
        "unidad_medida",
        "precio_publico",
        "precio_instalador",
        "precio_mayorista",
        "costo",
        "proveedor",
        "codigo_barras",
        "stock_deposito",
        "activo",
        "notas",
    ]
    ws.append(headers)
    style_header(ws, 1, len(headers))

    examples = [
        [
            "VS-1234",
            "Válvula de seguridad 1/2 termotanque",
            "Escorial",
            "Termotanques / Repuestos",
            "Unidades",
            18500,
            16650,
            15725,
            11200,
            "Distribuidor Escorial",
            "VS-1234",
            12,
            "si",
            "Ejemplo — borrar o reemplazar",
        ],
        [
            "TR-5678",
            "Termocouple universal cocina",
            "Argeflex",
            "Cocinas / Repuestos",
            "Unidades",
            9200,
            8280,
            7820,
            5600,
            "Importador Argeflex",
            "",
            25,
            "si",
            "",
        ],
        [
            "QU-9012",
            "Quemador auxiliar calefactor",
            "Eskabe",
            "Calefactores / Repuestos",
            "Unidades",
            45000,
            40500,
            38250,
            28000,
            "Eskabe SA",
            "QU-9012",
            3,
            "si",
            "Consultar portal web Eskabe",
        ],
    ]
    for row in examples:
        ws.append(row)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    activo_dv = DataValidation(type="list", formula1='"si,no"', allow_blank=True)
    activo_dv.error = "Usá si o no"
    activo_dv.errorTitle = "Valor inválido"
    ws.add_data_validation(activo_dv)
    activo_dv.add(f"M2:M50000")

    unidad_dv = DataValidation(
        type="list",
        formula1='"Unidades,kg,metros,litros,pares,cajas"',
        allow_blank=True,
    )
    ws.add_data_validation(unidad_dv)
    unidad_dv.add(f"E2:E50000")

    autosize_columns(ws)

    # --- Hoja instrucciones ---
    inst = wb.create_sheet("Instrucciones")
    inst["A1"] = "Maestro de productos — Servigas → Odoo 19"
    inst["A1"].font = Font(bold=True, size=14)

    lines = [
        "",
        "Cómo usar esta planilla",
        "1. Completá una fila por artículo. No dejar filas vacías entre productos.",
        "2. codigo_fabricante: obligatorio, único. Es la referencia principal en POS.",
        "3. descripcion: nombre legible del repuesto (aparece en POS y Factura Web).",
        "4. marca / categoria_producto: usar valores consistentes (misma ortografía).",
        "5. precio_*: números sin símbolo $; usar punto para decimales si aplica.",
        "6. stock_deposito: cantidad física en depósito al momento del go-live.",
        "7. activo: si = se importa; no = obsoleto (no borrar del Excel, marcar no).",
        "",
        "Mapeo columna → Odoo (import Inventario → Productos)",
        "codigo_fabricante  →  Referencia interna (default_code)",
        "descripcion        →  Nombre",
        "marca              →  Categoría de producto o campo Marca (según config)",
        "categoria_producto →  Categoría de producto",
        "unidad_medida      →  Unidad de medida",
        "precio_publico     →  Lista de precios Público",
        "precio_instalador  →  Lista de precios Instalador",
        "precio_mayorista   →  Lista de precios Mayorista",
        "costo              →  Costo",
        "proveedor          →  Proveedor (crear contactos antes del import)",
        "codigo_barras      →  Código de barras (opcional)",
        "stock_deposito     →  Ajuste de inventario (paso separado post-import)",
        "",
        "Reglas de limpieza antes de importar",
        "• Sin códigos duplicados.",
        "• Sin celdas combinadas.",
        "• Misma marca escrita igual en todas las filas (ej. Escorial, no ESCORIAL).",
        "• Revisar productos sin precio o sin costo.",
        "",
        "Pasos en Odoo",
        "1. Crear categorías de producto y listas de precio (Público, Instalador, Mayorista).",
        "2. Crear proveedores principales.",
        "3. Inventario → Productos → Favoritos → Importar → subir hoja Productos.",
        "4. Inventario → Ajustes → cargar stock_deposito por ubicación Depósito.",
        "",
        "AFIP / Factura Web: no incluido en esta etapa. POS sin factura fiscal por ahora.",
    ]
    for i, line in enumerate(lines, start=2):
        inst.cell(row=i, column=1, value=line)
        if line.startswith(("Cómo", "Mapeo", "Reglas", "Pasos")):
            inst.cell(row=i, column=1).font = Font(bold=True)
            inst.cell(row=i, column=1).fill = INSTR_FILL

    inst.column_dimensions["A"].width = 90

    # --- Hoja categorías sugeridas ---
    cats = wb.create_sheet("Categorias_sugeridas")
    cats.append(["categoria_producto", "ejemplos"])
    style_header(cats, 1, 2)
    sugeridas = [
        ("Termotanques / Repuestos", "Válvulas, ánodos, termostatos, resistencias"),
        ("Cocinas / Repuestos", "Quemadores, termocouples, perillas, tapas"),
        ("Calefactores / Repuestos", "Quemadores, pilas, ventiladores"),
        ("Anafes / Repuestos", "Mecheros, válvulas, flexibles"),
        ("Accesorios gas", "Flexibles, llaves, reguladores"),
        ("Repuestos genéricos", "Juntas, tornillería, selladores"),
    ]
    for cat, ej in sugeridas:
        cats.append([cat, ej])
    autosize_columns(cats)

    path = OUTPUT_DIR / "maestro_productos_odoo.xlsx"
    wb.save(path)
    print(f"Creado: {path}")


def create_planilla_puente() -> None:
    wb = Workbook()

    # --- Registro ventas ---
    ws = wb.active
    ws.title = "Registro_diario"

    headers = [
        "fecha",
        "hora",
        "n_ticket_odoo",
        "cliente",
        "cuit_dni",
        "condicion_iva",
        "total_odoo",
        "pide_factura",
        "n_factura_factura_web",
        "total_factura_web",
        "cuadra",
        "observaciones",
    ]
    ws.append(headers)
    style_header(ws, 1, len(headers))

    ws.append(
        [
            "2026-07-02",
            "10:30",
            "POS/00001",
            "Consumidor final",
            "",
            "Consumidor final",
            45200,
            "no",
            "",
            "",
            "si",
            "Solo ticket Odoo",
        ]
    )
    ws.append(
        [
            "2026-07-02",
            "11:15",
            "POS/00002",
            "García Instalaciones",
            "20-12345678-9",
            "Responsable Inscripto",
            128500,
            "si",
            "0001-00001234",
            128500,
            "si",
            "",
        ]
    )

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    pide_factura_dv = DataValidation(type="list", formula1='"si,no"', allow_blank=False)
    ws.add_data_validation(pide_factura_dv)
    pide_factura_dv.add("H2:H50000")

    cuadra_dv = DataValidation(type="list", formula1='"si,no,pendiente"', allow_blank=True)
    ws.add_data_validation(cuadra_dv)
    cuadra_dv.add("K2:K50000")

    condicion_dv = DataValidation(
        type="list",
        formula1='"Consumidor final,Monotributo,Responsable Inscripto,Exento"',
        allow_blank=True,
    )
    ws.add_data_validation(condicion_dv)
    condicion_dv.add("F2:F50000")

    autosize_columns(ws)

    # --- Cierre del día ---
    cierre = wb.create_sheet("Cierre_dia")
    cierre["A1"] = "Cierre diario — Odoo POS vs Factura Web"
    cierre["A1"].font = Font(bold=True, size=14)

    cierre_rows = [
        ("", ""),
        ("Fecha", ""),
        ("Responsable cierre", ""),
        ("", ""),
        ("Totales del día", "Importe"),
        ("Total ventas Odoo POS (sesión cerrada)", ""),
        ("Cantidad tickets Odoo", ""),
        ("Ventas con factura (filas pide_factura=si)", ""),
        ("Suma total_factura_web", ""),
        ("Ventas sin factura (tickets - con factura)", ""),
        ("", ""),
        ("Control", "Resultado"),
        ("¿Total Odoo = Factura Web + sin factura?", ""),
        ("¿Todas las filas con factura tienen n_factura?", ""),
        ("¿Alguna fila cuadra=no sin resolver?", ""),
        ("", ""),
        ("Notas del día", ""),
    ]
    for i, (label, value) in enumerate(cierre_rows, start=2):
        cierre.cell(row=i, column=1, value=label)
        cierre.cell(row=i, column=2, value=value)
        if label in {
            "Totales del día",
            "Control",
            "Cierre diario — Odoo POS vs Factura Web",
        } or label == "Fecha":
            cierre.cell(row=i, column=1).font = Font(bold=True)
            if label in {"Totales del día", "Control"}:
                cierre.cell(row=i, column=1).fill = INSTR_FILL
                cierre.cell(row=i, column=2).fill = INSTR_FILL

    cierre.column_dimensions["A"].width = 45
    cierre.column_dimensions["B"].width = 25

    # --- Instrucciones ---
    inst = wb.create_sheet("Instrucciones")
    inst["A1"] = "Planilla puente manual — Servigas"
    inst["A1"].font = Font(bold=True, size=14)

    lines = [
        "",
        "Regla de oro",
        "Siempre vender primero en Odoo POS. Factura Web solo si el cliente pide factura.",
        "",
        "Por cada venta con factura",
        "1. Cobrar en Odoo POS (mismos ítems y total).",
        "2. Emitir factura en Factura Web con los mismos datos.",
        "3. Anotar en Registro_diario: n_ticket_odoo + n_factura_factura_web.",
        "4. Verificar que total_odoo = total_factura_web → cuadra = si.",
        "",
        "Ventas sin factura",
        "Solo completar hasta total_odoo. pide_factura = no. Dejar vacío n_factura.",
        "",
        "Cierre del día (5-10 min)",
        "1. Cerrar sesión POS en Odoo y anotar total en hoja Cierre_dia.",
        "2. Contar filas del día en Registro_diario.",
        "3. Verificar que ventas con factura tienen número de Factura Web.",
        "4. Resolver filas con cuadra = no antes de irte.",
        "",
        "1 PC mostrador: Odoo en pestaña 1, Factura Web en pestaña 2.",
    ]
    for i, line in enumerate(lines, start=2):
        inst.cell(row=i, column=1, value=line)
        if line in {
            "Regla de oro",
            "Por cada venta con factura",
            "Ventas sin factura",
            "Cierre del día (5-10 min)",
        }:
            inst.cell(row=i, column=1).font = Font(bold=True)
            inst.cell(row=i, column=1).fill = INSTR_FILL

    inst.column_dimensions["A"].width = 85

    path = OUTPUT_DIR / "planilla_puente_factura_web.xlsx"
    wb.save(path)
    print(f"Creado: {path}")


if __name__ == "__main__":
    create_maestro_productos()
    create_planilla_puente()
