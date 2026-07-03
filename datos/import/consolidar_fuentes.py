"""Consolida listas de precios Servigas en maestro_productos_odoo.xlsx."""

from __future__ import annotations

import re
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from PyPDF2 import PdfReader

FUENTES_DIR = Path(r"C:\Users\mauri\OneDrive\Desktop\excelservigas")
OUTPUT_DIR = Path(__file__).parent

MAESTRO_COLUMNS = [
    "codigo_fabricante",
    "descripcion",
    "marca",
    "categoria_producto",
    "unidad_medida",
    "precio_publico",
    "precio_instalador",
    "precio_mayorista",
    "costo",
    "proveedor",
    "codigo_barras",
    "stock_deposito",
    "activo",
    "notas",
    "fuente_archivo",
]


def empty_row() -> dict:
    return {col: "" for col in MAESTRO_COLUMNS}


def normalize_codigo(value) -> str:
    if pd.isna(value):
        return ""
    text = str(value).strip()
    if text.endswith(".0") and text.replace(".", "", 1).isdigit():
        text = text[:-2]
    return text


def normalize_price(value) -> float | str:
    if pd.isna(value) or value == "":
        return ""
    if isinstance(value, (int, float)):
        return round(float(value), 2)
    text = str(value).strip()
    text = text.replace("$", "").replace(" ", "")
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        return round(float(text), 2)
    except ValueError:
        return ""


def make_product(**kwargs) -> dict:
    row = empty_row()
    row.update(kwargs)
    row["unidad_medida"] = row.get("unidad_medida") or "Unidades"
    row["activo"] = row.get("activo") or "si"
    return row


def parse_fercor(path: Path) -> list[dict]:
    df = pd.read_excel(path, header=None)
    rows = []
    for _, r in df.iterrows():
        codigo = normalize_codigo(r.iloc[0])
        if not codigo:
            continue
        desc = str(r.iloc[1]).strip() if pd.notna(r.iloc[1]) else ""
        if not desc or desc.lower() == "nan":
            continue
        precio = normalize_price(r.iloc[2])
        stock = r.iloc[3] if pd.notna(r.iloc[3]) else ""
        rows.append(
            make_product(
                codigo_fabricante=codigo,
                descripcion=desc,
                marca="Varios",
                categoria_producto="Repuestos / Inventario Fercor",
                precio_publico=precio,
                costo=precio,
                proveedor="Fercor",
                codigo_barras=codigo,
                stock_deposito=stock if stock != "" else "",
                notas="Inventario Fercor nov-2024. Precio y stock del archivo original.",
                fuente_archivo=path.name,
            )
        )
    return rows


def parse_brogas(path: Path) -> list[dict]:
    df = pd.read_excel(path, sheet_name="LISTA REPUESTOS SERVICE.", header=2)
    df.columns = ["origen", "codigo", "descripcion", "precio"]
    rows = []
    current_section = "Repuestos Brogas"
    for _, r in df.iterrows():
        origen = str(r["origen"]).strip() if pd.notna(r["origen"]) else ""
        codigo = normalize_codigo(r["codigo"])
        desc = str(r["descripcion"]).strip() if pd.notna(r["descripcion"]) else ""
        if not codigo and desc and origen.upper() not in {"ORIGEN", "NAN"}:
            current_section = desc
            continue
        if not codigo or not desc:
            continue
        if codigo.upper() == "CODIGO":
            continue
        rows.append(
            make_product(
                codigo_fabricante=codigo,
                descripcion=desc,
                marca="Brogas",
                categoria_producto=current_section,
                precio_publico=normalize_price(r["precio"]),
                proveedor="Brogas",
                codigo_barras=codigo,
                notas=f"Origen lista: {origen}" if origen else "",
                fuente_archivo=path.name,
            )
        )
    return rows


def parse_carr_sur(path: Path) -> list[dict]:
    df = pd.read_excel(path, sheet_name="PRECIOS", header=9)
    df.columns = ["codigo", "descripcion", "moneda", "precio", "iva"]
    rows = []
    for _, r in df.iterrows():
        codigo = normalize_codigo(r["codigo"])
        desc = str(r["descripcion"]).strip() if pd.notna(r["descripcion"]) else ""
        if not codigo or not desc or codigo.lower() == "codigo":
            continue
        iva = normalize_price(r["iva"])
        notas = "Lista Carr-Sur"
        if iva != "":
            notas += f" | IVA {iva}%"
        rows.append(
            make_product(
                codigo_fabricante=codigo,
                descripcion=desc,
                marca="Varios",
                categoria_producto="Repuestos gas / Carr-Sur",
                precio_publico=normalize_price(r["precio"]),
                proveedor="Fernando Miguel Carreras (Carr-Sur)",
                codigo_barras=codigo,
                notas=notas,
                fuente_archivo=path.name,
            )
        )
    return rows


def lista_vigente_codigo(marca: str, codigo: str) -> str:
    """Gran Ferretero usa códigos internos (1, 2, 3). Prefijamos con marca."""
    marca_slug = re.sub(r"[^A-Za-z0-9]+", "", marca.upper())[:12] or "GF"
    if codigo.isdigit() or len(codigo) <= 4:
        return f"{marca_slug}-{codigo}"
    return codigo


def parse_lista_vigente(path: Path) -> list[dict]:
    df = pd.read_excel(path, sheet_name="Hoja1", header=None)
    data = df.iloc[5:, 1:].copy()
    data.columns = [
        "marca",
        "codigo",
        "descripcion",
        "rubro",
        "subrubro",
        "pvp",
        "pcb",
        "compra_minima",
        "cantidad",
        "_extra",
    ]
    rows = []
    for _, r in data.iterrows():
        codigo_raw = normalize_codigo(r["codigo"])
        desc = str(r["descripcion"]).strip() if pd.notna(r["descripcion"]) else ""
        if not codigo_raw or not desc:
            continue
        marca = str(r["marca"]).strip() if pd.notna(r["marca"]) else "Varios"
        codigo = lista_vigente_codigo(marca, codigo_raw)
        rubro = str(r["rubro"]).strip() if pd.notna(r["rubro"]) else ""
        subrubro = str(r["subrubro"]).strip() if pd.notna(r["subrubro"]) else ""
        categoria = " / ".join(x for x in [rubro, subrubro] if x)
        rows.append(
            make_product(
                codigo_fabricante=codigo,
                descripcion=desc,
                marca=marca,
                categoria_producto=categoria or "Gran Ferretero",
                precio_publico=normalize_price(r["pvp"]),
                proveedor="Gran Ferretero",
                codigo_barras=codigo,
                notas=f"Lista vigente junio 2026 | codigo original GF: {codigo_raw}",
                fuente_archivo=path.name,
            )
        )
    return rows


def parse_eskabe_sheet(df: pd.DataFrame, sheet: str, categoria: str) -> list[dict]:
    rows = []
    cod_col = df.columns[0]
    price_col = [c for c in df.columns if "IVA" in str(c).upper() or "VALOR" in str(c).upper()][-1]
    for _, r in df.iterrows():
        codigo = normalize_codigo(r[cod_col])
        if not codigo or codigo.upper() == "CODIGO":
            continue
        desc = str(r.iloc[1]).strip() if pd.notna(r.iloc[1]) else ""
        if not desc:
            continue
        extra = []
        for col in df.columns[2:-1]:
            val = r[col]
            if pd.notna(val) and str(val).strip():
                extra.append(f"{col}={val}")
        rows.append(
            make_product(
                codigo_fabricante=codigo,
                descripcion=desc,
                marca="Eskabe",
                categoria_producto=categoria,
                precio_publico=normalize_price(r[price_col]),
                proveedor="Eskabe",
                codigo_barras=codigo,
                notas="Precio sin IVA | " + " | ".join(extra[:4]),
                fuente_archivo="REPUESTOS JUNIO 2025.xlsx",
            )
        )
    return rows


def parse_eskabe(path: Path) -> list[dict]:
    mapping = {
        "AGUA CALIENTE": "Termotanques / Agua caliente",
        "CALEFACCION": "Calefactores / Repuestos",
        "COCCION": "Cocinas / Repuestos",
    }
    rows = []
    seen_codes = set()
    for sheet, categoria in mapping.items():
        df = pd.read_excel(path, sheet_name=sheet)
        for row in parse_eskabe_sheet(df, sheet, categoria):
            if row["codigo_fabricante"] in seen_codes:
                continue
            seen_codes.add(row["codigo_fabricante"])
            rows.append(row)
    return rows


def parse_longvie_pdf(path: Path) -> list[dict]:
    reader = PdfReader(str(path))
    text = "\n".join(page.extract_text() or "" for page in reader.pages)
    lines = text.splitlines()
    rows = []
    pattern = re.compile(r"^([A-Z]{2}[\w.]+)\s+(.+?)\s+\$\s*([\d.,]+)\s*$")
    for line in lines:
        line = line.strip()
        if not line or line.startswith("Lista de Precios") or "Código de artículo" in line:
            continue
        match = pattern.match(line)
        if not match:
            continue
        codigo, desc, precio = match.groups()
        rows.append(
            make_product(
                codigo_fabricante=codigo,
                descripcion=desc.strip(),
                marca="Longvie",
                categoria_producto="Repuestos Longvie",
                precio_publico=normalize_price(precio),
                proveedor="Longvie",
                codigo_barras=codigo,
                notas="Lista repuestos junio 2026 PDF | precio sin IVA",
                fuente_archivo=path.name,
            )
        )
    return rows


def deduplicate_within_source(rows: list[dict]) -> list[dict]:
    seen: dict[tuple[str, str], dict] = {}
    for row in rows:
        key = (row["fuente_archivo"], row["codigo_fabricante"])
        if key not in seen:
            seen[key] = row
    return list(seen.values())


def deduplicate(rows: list[dict]) -> tuple[list[dict], pd.DataFrame]:
  priority = {
      "20241128Fercor.xlsx": 100,
      "REPUESTOS JUNIO 2025.xlsx": 90,
      "Lista de Precios Repuestos Junio  2026.pdf": 85,
      "BROGAS LISTA REPUESTOS- 17-03-2025.xlsx": 80,
      "CARR-SUR-Precios(17).xls": 70,
      "CARR-SUR-Precios(16).xls": 60,
      "LISTA VIGENTE JUNIO 2026.xlsx": 50,
  }

  by_code: dict[str, list[dict]] = {}
  for row in rows:
      code = row["codigo_fabricante"]
      by_code.setdefault(code, []).append(row)

  final = []
  dup_report = []
  for code, items in sorted(by_code.items()):
      if len(items) == 1:
          final.append(items[0])
          continue
      items_sorted = sorted(
          items,
          key=lambda x: priority.get(x["fuente_archivo"], 0),
          reverse=True,
      )
      chosen = items_sorted[0]
      final.append(chosen)
      dup_report.append(
          {
              "codigo_fabricante": code,
              "elegido": chosen["fuente_archivo"],
              "descartados": ", ".join(
                  f"{i['fuente_archivo']} ({i['marca']})" for i in items_sorted[1:]
              ),
              "descripcion_elegida": chosen["descripcion"],
          }
      )

  return final, pd.DataFrame(dup_report)


def style_workbook(path: Path) -> None:
    wb = load_workbook(path)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for ws in wb.worksheets:
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
    wb.save(path)


def main() -> None:
    sources = [
        ("fercor", FUENTES_DIR / "20241128Fercor.xlsx", parse_fercor),
        ("brogas", FUENTES_DIR / "BROGAS LISTA REPUESTOS- 17-03-2025.xlsx", parse_brogas),
        ("carr17", FUENTES_DIR / "CARR-SUR-Precios(17).xls", parse_carr_sur),
        ("lista_vigente", FUENTES_DIR / "LISTA VIGENTE JUNIO 2026.xlsx", parse_lista_vigente),
        ("eskabe", FUENTES_DIR / "REPUESTOS JUNIO 2025.xlsx", parse_eskabe),
        ("longvie_pdf", FUENTES_DIR / "Lista de Precios Repuestos Junio  2026.pdf", parse_longvie_pdf),
    ]

    all_rows: list[dict] = []
    stats = []

    for key, path, parser in sources:
        if not path.exists():
            stats.append({"fuente": key, "archivo": path.name, "filas": 0, "estado": "NO ENCONTRADO"})
            continue
        parsed = deduplicate_within_source(parser(path))
        all_rows.extend(parsed)
        stats.append({"fuente": key, "archivo": path.name, "filas": len(parsed), "estado": "OK"})

    consolidated, duplicates = deduplicate(all_rows)
    df = pd.DataFrame(consolidated)
    export_cols = [c for c in MAESTRO_COLUMNS if c != "fuente_archivo"]
    df_export = df[export_cols + ["fuente_archivo"]].sort_values(["marca", "codigo_fabricante"])

    out_maestro = OUTPUT_DIR / "maestro_productos_consolidado.xlsx"
    out_report = OUTPUT_DIR / "reporte_consolidacion.xlsx"

    with pd.ExcelWriter(out_maestro, engine="openpyxl") as writer:
        df_export.to_excel(writer, sheet_name="Productos", index=False)
        pd.DataFrame(stats).to_excel(writer, sheet_name="Resumen_fuentes", index=False)
        if not duplicates.empty:
            duplicates.to_excel(writer, sheet_name="Duplicados_resueltos", index=False)

    with pd.ExcelWriter(out_report, engine="openpyxl") as writer:
        pd.DataFrame(stats).to_excel(writer, sheet_name="Fuentes", index=False)
        df["fuente_archivo"].value_counts().reset_index().to_excel(
            writer, sheet_name="Por_archivo", index=False
        )
        df["marca"].value_counts().reset_index().to_excel(
            writer, sheet_name="Por_marca", index=False
        )
        if not duplicates.empty:
            duplicates.to_excel(writer, sheet_name="Duplicados", index=False)
        sin_precio = df[df["precio_publico"].isin(["", None])]
        sin_precio.head(500).to_excel(writer, sheet_name="Sin_precio", index=False)

    style_workbook(out_maestro)
    style_workbook(out_report)

    print(f"Total filas parseadas: {len(all_rows)}")
    print(f"Total consolidado (unicos): {len(consolidated)}")
    print(f"Duplicados resueltos: {len(duplicates)}")
    for s in stats:
        print(f"  {s['archivo']}: {s['filas']} ({s['estado']})")
    print(f"\nGenerado: {out_maestro}")
    print(f"Generado: {out_report}")


if __name__ == "__main__":
    main()
